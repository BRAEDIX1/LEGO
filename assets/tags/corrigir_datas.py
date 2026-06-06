"""
corrigir_datas.py
-----------------
Corrige as colunas 'data' (texto dd/mm/yyyy) e 'hora' (texto HH:MM:SS)
de resultado_consolidado.xlsx para tipos reais de data e hora no Excel.

Uso:
    python corrigir_datas.py                              # usa caminhos padrão
    python corrigir_datas.py entrada.xlsx                 # define arquivo de entrada
    python corrigir_datas.py entrada.xlsx saida.xlsx      # define entrada e saída
"""

import sys
import os
import pandas as pd
from openpyxl.utils import get_column_letter


def corrigir(entrada: str, saida: str) -> None:
    if not os.path.exists(entrada):
        print(f"[ERRO] Arquivo não encontrado: {entrada}")
        sys.exit(1)

    print(f"Lendo: {entrada}")
    df = pd.read_excel(entrada, dtype=str)

    # --- corrigir coluna 'data' ---
    if 'data' not in df.columns:
        print("[AVISO] Coluna 'data' não encontrada. Pulando.")
    else:
        antes = df['data'].dtype
        df['data'] = pd.to_datetime(df['data'], format='%d/%m/%Y', errors='coerce')
        nulos = df['data'].isna().sum()
        print(f"  data   : {antes} → datetime64  |  valores inválidos: {nulos}")

    # --- corrigir coluna 'hora' ---
    if 'hora' not in df.columns:
        print("[AVISO] Coluna 'hora' não encontrada. Pulando.")
    else:
        antes = df['hora'].dtype
        df['hora'] = pd.to_timedelta(df['hora'], errors='coerce')
        nulos = df['hora'].isna().sum()
        print(f"  hora   : {antes} → timedelta64  |  valores inválidos: {nulos}")

    # --- salvar ---
    print(f"Salvando: {saida}")
    with pd.ExcelWriter(saida, engine='openpyxl', datetime_format='DD/MM/YYYY') as writer:
        df.to_excel(writer, index=False, sheet_name='Consolidado')
        ws = writer.sheets['Consolidado']

        # aplicar formatos de exibição célula a célula
        col_data = col_hora = None
        for i, col in enumerate(df.columns, 1):
            if col == 'data':
                col_data = get_column_letter(i)
            elif col == 'hora':
                col_hora = get_column_letter(i)

        for row in range(2, ws.max_row + 1):
            if col_data:
                ws[f'{col_data}{row}'].number_format = 'DD/MM/YYYY'
            if col_hora:
                ws[f'{col_hora}{row}'].number_format = 'HH:MM:SS'

        # ajustar largura das colunas
        for col in ws.columns:
            largura = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(largura + 2, 30)

    print("Concluído!")
    if 'data' in df.columns:
        print(f"  Data mais recente : {df['data'].max().strftime('%d/%m/%Y')}")
        print(f"  Data mais antiga  : {df['data'].min().strftime('%d/%m/%Y')}")


if __name__ == '__main__':
    args = sys.argv[1:]
    entrada_padrao = 'resultado_consolidado.xlsx'
    saida_padrao   = 'resultado_consolidado_corrigido.xlsx'

    if len(args) == 0:
        _entrada, _saida = entrada_padrao, saida_padrao
    elif len(args) == 1:
        _entrada = args[0]
        base, ext = os.path.splitext(_entrada)
        _saida = f"{base}_corrigido{ext}"
    else:
        _entrada, _saida = args[0], args[1]

    corrigir(_entrada, _saida)
