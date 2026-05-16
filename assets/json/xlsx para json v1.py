"""
xlsx_para_json.py
-----------------
Converte firestore_dump.xlsx para firestore_dump.json
no mesmo formato esperado pelo app LEGO.

Uso:
    python xlsx_para_json.py
    python xlsx_para_json.py caminho/para/firestore_dump.xlsx
    python xlsx_para_json.py entrada.xlsx saida.json

Dependências:
    pip install openpyxl
"""

import sys
import json
from pathlib import Path
from datetime import datetime
import openpyxl


# ---------------------------------------------------------------------------
# Configuração
# ---------------------------------------------------------------------------
DEFAULT_INPUT  = "firestore_dump.xlsx"
DEFAULT_OUTPUT = "firestore_dump.json"
TAG_DIGITS     = 9  # comprimento esperado de cada TAG


# ---------------------------------------------------------------------------
# Leitura das abas
# ---------------------------------------------------------------------------
def ler_gases(ws) -> dict:
    """
    Aba 'gases': colunas codigo | descricao | volume | unidade
    Chave do dict = str(codigo)
    """
    gases = {}
    cabecalho = None

    for row in ws.iter_rows(values_only=True):
        if cabecalho is None:
            cabecalho = [str(c).strip() for c in row]
            continue

        row = dict(zip(cabecalho, row))
        codigo = row.get("codigo")

        if codigo is None:
            continue

        codigo = str(int(codigo)) if isinstance(codigo, float) else str(codigo).strip()

        entrada = {"descricao": str(row.get("descricao") or "").strip()}

        volume  = row.get("volume")
        unidade = str(row.get("unidade") or "").strip()

        if unidade:
            entrada["unidade"] = unidade

        if volume is not None and str(volume).strip() not in ("", "None"):
            try:
                v = float(str(volume).replace(",", "."))
                entrada["volume"] = int(v) if v == int(v) else v
            except ValueError:
                pass

        gases[codigo] = entrada

    return gases


def ler_barras(ws) -> tuple:
    """
    Aba 'barras': colunas tag | codigo
    Chave do dict = str(tag)

    Retorna (barras, suspeitos) onde suspeitos são tags com menos de
    TAG_DIGITS dígitos — sinal de possível perda de zero à esquerda.
    """
    barras    = {}
    suspeitos = []
    cabecalho = None

    for row in ws.iter_rows(values_only=True):
        if cabecalho is None:
            cabecalho = [str(c).strip() for c in row]
            continue

        row = dict(zip(cabecalho, row))
        tag    = row.get("tag")
        codigo = row.get("codigo")

        if tag is None or codigo is None:
            continue

        tag    = str(tag).strip()
        codigo = str(int(codigo)) if isinstance(codigo, float) else str(codigo).strip()

        # Alerta se tag numérica tiver menos dígitos que o esperado
        if tag.isdigit() and len(tag) < TAG_DIGITS:
            suspeitos.append((tag, codigo))

        barras[tag] = {"codigo": codigo}

    return barras, suspeitos


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    args = sys.argv[1:]

    if len(args) >= 2:
        entrada = Path(args[0])
        saida   = Path(args[1])
    elif len(args) == 1:
        entrada = Path(args[0])
        saida   = entrada.with_suffix(".json")
    else:
        entrada = Path(DEFAULT_INPUT)
        saida   = Path(DEFAULT_OUTPUT)

    if not entrada.exists():
        print(f"Arquivo não encontrado: {entrada}")
        sys.exit(1)

    print(f"Lendo: {entrada}")
    wb = openpyxl.load_workbook(entrada, read_only=True)

    if "gases" not in wb.sheetnames:
        print("ERRO: aba 'gases' não encontrada.")
        sys.exit(1)

    if "barras" not in wb.sheetnames:
        print("ERRO: aba 'barras' não encontrada.")
        sys.exit(1)

    gases             = ler_gases(wb["gases"])
    barras, suspeitos = ler_barras(wb["barras"])

    versao = datetime.now().strftime("%Y.%m.%d-%H:%M")

    resultado = {
        "version": versao,
        "gases":   gases,
        "barras":  barras,
    }

    with open(saida, "w", encoding="utf-8") as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2)

    print(f"Gases:  {len(gases):,} registros")
    print(f"Barras: {len(barras):,} registros")
    print(f"Salvo:  {saida}")

    # Alerta de integridade
    if suspeitos:
        print(f"\nAVISO: {len(suspeitos):,} tag(s) com menos de {TAG_DIGITS} digitos — possivel perda de zero a esquerda no Excel:")
        for tag, cod in suspeitos[:20]:
            print(f"  TAG '{tag}' (codigo {cod}) — esperado: '{tag.zfill(TAG_DIGITS)}'")
        if len(suspeitos) > 20:
            print(f"  ... e mais {len(suspeitos) - 20} ocorrencias.")
        print("\nSolucao: formate a coluna 'tag' no xlsx como Texto antes de salvar.")
    else:
        print(f"\nIntegridade OK — todas as tags tem {TAG_DIGITS} digitos.")


if __name__ == "__main__":
    main()
